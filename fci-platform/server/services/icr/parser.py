"""
Spreadsheet parsing for .xlsx and .csv files.
Handles both regular single-sheet spreadsheets and
multi-sheet UOL (User Operation Log) workbooks.
"""

import io
import csv
import openpyxl
from .utils import safe_str, safe_float, safe_int


def parse_uploaded_file(filename, file_bytes):
    """
    Parse an uploaded file into structured data.

    Returns a list of parsed results. Most files produce a single
    result, but multi-sheet Excel workbooks produce one per sheet.
    UOL workbooks produce a single UOL result dict (with is_uol=True).

    Args:
        filename: Original filename
        file_bytes: Raw file content as bytes
    """
    lower = filename.lower()

    if lower.endswith('.csv'):
        return [_parse_csv(filename, file_bytes)]
    elif lower.endswith(('.xlsx', '.xls')):
        return _parse_excel(filename, file_bytes)
    else:
        return [{
            'filename': filename,
            'error': 'Unsupported file format',
            'headers': [],
            'rows': [],
        }]


def _parse_csv(filename, file_bytes):
    """Parse a CSV file into headers + row dicts."""
    try:
        text = file_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = file_bytes.decode('latin-1')

    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)

    if not all_rows:
        return {'filename': filename, 'headers': [], 'rows': []}

    headers = [safe_str(h) for h in all_rows[0]]
    data_rows = []
    for row in all_rows[1:]:
        obj = {}
        for j, header in enumerate(headers):
            obj[header] = row[j] if j < len(row) else ''
        data_rows.append(obj)

    return {
        'filename': filename,
        'headers': headers,
        'rows': data_rows,
    }


def _parse_excel(filename, file_bytes):
    """Parse an Excel file. Detects UOL workbooks vs regular spreadsheets."""
    # First pass: read_only mode to check if it's a UOL workbook
    wb = openpyxl.load_workbook(
        io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )

    try:
        is_uol = _is_uol_workbook(wb)
    finally:
        wb.close()

    if is_uol:
        # Re-open WITHOUT read_only — UOL sheets have formatting
        # that causes read_only mode to miss rows
        wb_full = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
        )
        try:
            return [_parse_uol_workbook(filename, wb_full)]
        finally:
            wb_full.close()
    else:
        # Regular spreadsheets — parse all sheets
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
        try:
            return _parse_single_sheet(filename, wb)
        finally:
            wb.close()


def _is_uol_workbook(wb):
    """Check if workbook is a UOL file by looking for expected sheet names."""
    names = [n.lower().strip() for n in wb.sheetnames]
    return 'customer information' in names and len(wb.sheetnames) >= 3


def _parse_single_sheet(filename, wb):
    """
    Parse all sheets of a non-UOL workbook.

    Returns a list of parsed results — one per sheet that has data.
    Single-sheet workbooks return a one-element list.
    """
    results = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        if not all_rows:
            continue

        headers = [safe_str(cell) for cell in all_rows[0]]
        if not any(headers):
            continue

        data_rows = []
        for row in all_rows[1:]:
            obj = {}
            for j, header in enumerate(headers):
                obj[header] = row[j] if j < len(row) else ''
            data_rows.append(obj)

        sheet_label = f'{filename} [{sheet_name}]' if len(wb.sheetnames) > 1 else filename
        results.append({
            'filename': sheet_label,
            'headers': headers,
            'rows': data_rows,
        })

    if not results:
        return [{'filename': filename, 'headers': [], 'rows': []}]

    return results


def _parse_uol_workbook(filename, wb):
    """
    Parse a UOL (User Operation Log) workbook.
    Extracts Customer Information, Fiat Deposit History,
    Fiat Withdrawal History, Withdrawal History (crypto),
    and Deposit History (crypto) tabs.
    """
    result = {
        'is_uol': True,
        'filename': filename,
        'sheet_names': wb.sheetnames,
        'customer_info': None,
        'fiat_deposits': [],
        'fiat_withdrawals': [],
        'crypto_withdrawals': [],
        'crypto_deposits': [],
        'attempted_withdrawals': [],
        'binance_pay': [],
        'p2p_transactions': [],
    }

    sheet_map = {}
    for name in wb.sheetnames:
        sheet_map[name.lower().strip()] = name


    if 'customer information' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['customer information']])
        result['customer_info'] = _parse_uol_customer_info(raw)

    if 'fiat deposit history' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['fiat deposit history']])
        result['fiat_deposits'] = _parse_uol_fiat_tab(raw)

    if 'fiat withdrawal history' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['fiat withdrawal history']])
        result['fiat_withdrawals'] = _parse_uol_fiat_tab(raw)

    if 'withdrawal history' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['withdrawal history']])
        result['crypto_withdrawals'] = _parse_uol_crypto_withdrawal_tab(raw)

    if 'deposit history' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['deposit history']])
        result['crypto_deposits'] = _parse_uol_crypto_deposit_tab(raw)

    if 'attempted withdrawal history' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['attempted withdrawal history']])
        result['attempted_withdrawals'] = _parse_uol_attempted_withdrawal_tab(raw)

    if 'binance pay' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['binance pay']])
        result['binance_pay'] = _parse_uol_binance_pay_tab(raw)

    if 'p2p' in sheet_map:
        raw = _sheet_to_raw_rows(wb[sheet_map['p2p']])
        result['p2p_transactions'] = _parse_uol_p2p_tab(raw)

    return result


def _sheet_to_raw_rows(ws):
    """Convert a worksheet to a list of lists (raw rows)."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _parse_uol_customer_info(raw_rows):
    """
    Parse the Customer Information tab from a UOL workbook.
    This tab has a header row followed by a single data row,
    but the header row position varies.
    """
    if not raw_rows:
        return None

    header_row_idx = -1
    header_map = {}

    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_user_id = False
        has_first_name = False
        for j, cell in enumerate(row):
            cell_val = safe_str(cell).lower()
            if cell_val == 'user id':
                has_user_id = True
            if cell_val == 'first name':
                has_first_name = True
        if has_user_id and has_first_name:
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return None

    # Find the first real data row after the header
    data_row = None
    skip_labels = {'api information', 'basic information'}
    for d in range(header_row_idx + 1, len(raw_rows)):
        candidate = raw_rows[d]
        if not candidate:
            continue
        first_cell = safe_str(candidate[0])
        if not first_cell:
            continue
        if first_cell.lower() in skip_labels:
            continue
        data_row = candidate
        break

    if data_row is None:
        return None

    def get_val(possible_keys):
        for key in possible_keys:
            idx = header_map.get(key.lower())
            if idx is not None and idx < len(data_row):
                return safe_str(data_row[idx])
        return ''

    first_name = get_val(['First Name'])
    last_name = get_val(['Last Name'])

    return {
        'user_id': get_val(['User ID']),
        'email': get_val(['Email']),
        'mobile': get_val(['Mobile']),
        'registration_time': get_val(['Registration time']),
        'first_name': first_name,
        'last_name': last_name,
        'full_name': (first_name + ' ' + last_name).strip(),
        'nationality': get_val(['User nationality']),
        'id_number': get_val(['User ID number']),
        'auth_type': get_val(['User authentication type']),
        'auth_time': get_val(['User authentication time']),
        'status': get_val(['Status']),
        'two_fa': get_val(['2FA']),
        'sms': get_val(['SMS']),
        'tax_id': get_val(['Tax ID']),
        'tnc': get_val(['TnC']),
        'tnc_sign_date': get_val(['TnC Sign Date']),
    }


def _parse_uol_fiat_tab(raw_rows):
    """
    Parse a fiat transaction tab (deposits or withdrawals) from UOL.
    Handles variable header positions and column names.
    """
    if not raw_rows:
        return []

    header_row_idx = -1
    header_map = {}

    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_status = False
        has_order_create = False
        for j, cell in enumerate(row):
            cell_val = safe_str(cell).lower()
            if cell_val in ('status name', 'status'):
                has_status = True
            if cell_val == 'order create time':
                has_order_create = True
        if has_status and has_order_create:
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return []

    def col_idx(possible_names):
        for name in possible_names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return -1

    i_status = col_idx(['Status Name', 'Status'])
    i_date = col_idx(['Order Create Time'])
    i_currency = col_idx(['Currency'])
    i_amount = col_idx(['Gross Amount'])
    i_amount_usd = col_idx(['Gross Amount Usd', 'Gross Amount USD'])
    i_channel = col_idx(['Channel Name'])
    i_method = col_idx(['Transaction Method'])
    i_card_bin = col_idx(['Card Bin', 'Card BIN'])
    i_card_last4 = col_idx(['Card Last 4 Digital', 'Card Last 4 Digits'])
    i_order_id = col_idx(['Order Id', 'Order ID'])
    i_bank_name = col_idx(['Issuer Bank Name'])
    i_bank_country = col_idx(['Issuer Bank Country'])
    i_viban = col_idx(['Viban', 'VIBAN'])
    i_iban = col_idx(['Iban', 'IBAN'])
    i_account_number = col_idx(['Account Number'])

    def cell_val(row, idx):
        if idx < 0 or idx >= len(row):
            return ''
        return row[idx]

    results = []
    for r in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[r]
        if not row:
            continue
        status_val = safe_str(cell_val(row, i_status)) if i_status >= 0 else ''
        if not status_val:
            continue
        results.append({
            'status': status_val,
            'date': safe_str(cell_val(row, i_date)),
            'currency': safe_str(cell_val(row, i_currency)),
            'amount': safe_float(cell_val(row, i_amount)),
            'amount_usd': safe_float(cell_val(row, i_amount_usd)),
            'channel': safe_str(cell_val(row, i_channel)),
            'method': safe_str(cell_val(row, i_method)),
            'card_bin': safe_str(cell_val(row, i_card_bin)),
            'card_last4': safe_str(cell_val(row, i_card_last4)),
            'order_id': safe_str(cell_val(row, i_order_id)),
            'bank_name': safe_str(cell_val(row, i_bank_name)),
            'bank_country': safe_str(cell_val(row, i_bank_country)),
            'viban': safe_str(cell_val(row, i_viban)),
            'iban': safe_str(cell_val(row, i_iban)),
            'account_number': safe_str(cell_val(row, i_account_number)),
        })

    return results


def _parse_uol_crypto_withdrawal_tab(raw_rows):
    """
    Parse the Withdrawal History (crypto) tab from a UOL workbook.
    These are crypto withdrawals TO external addresses.
    """
    if not raw_rows:
        return []

    header_row_idx = -1
    header_map = {}

    # Identify header row by looking for key columns
    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_destination = False
        has_currency = False
        for j, cell in enumerate(row):
            cell_val = safe_str(cell).lower()
            if cell_val == 'destination address':
                has_destination = True
            if cell_val == 'currency':
                has_currency = True
        if has_destination and has_currency:
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return []

    def col_idx(possible_names):
        for name in possible_names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return -1

    i_user_id = col_idx(['User ID'])
    i_currency = col_idx(['Currency'])
    i_amount = col_idx(['Amount'])
    i_account_type = col_idx(['Account Type'])
    i_usdt = col_idx(['USDT'])
    i_destination = col_idx(['Destination Address'])
    i_label = col_idx(['Label/Tag/Memo', 'Label', 'Tag', 'Memo'])
    i_txid = col_idx(['txId', 'TxId', 'TXID', 'txID'])
    i_apply_time = col_idx(['Apply Time'])
    i_status = col_idx(['Status'])
    i_network = col_idx(['Network'])
    i_counterparty = col_idx(['CounterParty ID', 'Counterparty ID'])

    def cell_val(row, idx):
        if idx < 0 or idx >= len(row):
            return ''
        return row[idx]

    results = []
    for r in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[r]
        if not row:
            continue
        dest = safe_str(cell_val(row, i_destination))
        currency = safe_str(cell_val(row, i_currency))
        # Skip empty rows — require at least destination or currency
        if not dest and not currency:
            continue
        results.append({
            'user_id': safe_str(cell_val(row, i_user_id)),
            'currency': currency,
            'amount': safe_float(cell_val(row, i_amount)),
            'account_type': safe_str(cell_val(row, i_account_type)),
            'usdt_value': safe_float(cell_val(row, i_usdt)),
            'destination_address': dest,
            'label': safe_str(cell_val(row, i_label)),
            'tx_id': safe_str(cell_val(row, i_txid)),
            'apply_time': safe_str(cell_val(row, i_apply_time)),
            'status': safe_str(cell_val(row, i_status)),
            'network': safe_str(cell_val(row, i_network)),
            'counterparty_id': safe_str(cell_val(row, i_counterparty)),
        })

    return results


def _parse_uol_crypto_deposit_tab(raw_rows):
    """
    Parse the Deposit History (crypto) tab from a UOL workbook.
    These are crypto deposits FROM external addresses.
    """
    if not raw_rows:
        return []

    header_row_idx = -1
    header_map = {}

    # Identify header row by looking for key columns
    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_source = False
        has_deposit_addr = False
        has_currency = False
        for j, cell in enumerate(row):
            cell_val = safe_str(cell).lower()
            if cell_val == 'source address':
                has_source = True
            if cell_val == 'deposit address':
                has_deposit_addr = True
            if cell_val == 'currency':
                has_currency = True
        # Match if we have currency AND at least one address column
        if has_currency and (has_source or has_deposit_addr):
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return []

    def col_idx(possible_names):
        for name in possible_names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return -1

    i_user_id = col_idx(['User ID'])
    i_currency = col_idx(['Currency'])
    i_amount = col_idx(['Amount'])
    i_account_type = col_idx(['Account Type'])
    i_usdt = col_idx(['USDT'])
    i_deposit_addr = col_idx(['Deposit Address'])
    i_label = col_idx(['Label/Tag/Memo', 'Label', 'Tag', 'Memo'])
    i_source_addr = col_idx(['Source Address'])
    i_txid = col_idx(['TXID', 'txId', 'TxId', 'txID'])
    i_create_time = col_idx(['Create Time'])
    i_status = col_idx(['Status'])
    i_network = col_idx(['Network'])
    i_counterparty = col_idx(['CounterParty ID', 'Counterparty ID'])

    def cell_val(row, idx):
        if idx < 0 or idx >= len(row):
            return ''
        return row[idx]

    results = []
    for r in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[r]
        if not row:
            continue
        source = safe_str(cell_val(row, i_source_addr))
        deposit = safe_str(cell_val(row, i_deposit_addr))
        currency = safe_str(cell_val(row, i_currency))
        # Skip empty rows
        if not currency and not source and not deposit:
            continue
        results.append({
            'user_id': safe_str(cell_val(row, i_user_id)),
            'currency': currency,
            'amount': safe_float(cell_val(row, i_amount)),
            'account_type': safe_str(cell_val(row, i_account_type)),
            'usdt_value': safe_float(cell_val(row, i_usdt)),
            'deposit_address': deposit,
            'label': safe_str(cell_val(row, i_label)),
            'source_address': source,
            'tx_id': safe_str(cell_val(row, i_txid)),
            'create_time': safe_str(cell_val(row, i_create_time)),
            'status': safe_str(cell_val(row, i_status)),
            'network': safe_str(cell_val(row, i_network)),
            'counterparty_id': safe_str(cell_val(row, i_counterparty)),
        })

    return results


def _parse_uol_attempted_withdrawal_tab(raw_rows):
    """
    Parse the Attempted Withdrawal History tab from a UOL workbook.
    Columns: User ID, Asset, Network, Amount, Address, Date,
    Business Type, Source, DecisionCode, USDT Equivalent.
    """
    if not raw_rows:
        return []

    header_row_idx = -1
    header_map = {}

    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_address = False
        has_asset = False
        has_decision = False
        for j, cell in enumerate(row):
            cell_val = safe_str(cell).lower()
            if cell_val == 'address':
                has_address = True
            if cell_val == 'asset':
                has_asset = True
            if cell_val == 'decisioncode':
                has_decision = True
        if has_address and has_asset and has_decision:
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return []

    def col_idx(possible_names):
        for name in possible_names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return -1

    i_user_id = col_idx(['User ID'])
    i_asset = col_idx(['Asset'])
    i_network = col_idx(['Network'])
    i_amount = col_idx(['Amount'])
    i_address = col_idx(['Address'])
    i_date = col_idx(['Date'])
    i_business_type = col_idx(['Business Type'])
    i_source = col_idx(['Source'])
    i_usdt = col_idx(['USDT Equivalent'])

    def cell_val(row, idx):
        if idx < 0 or idx >= len(row):
            return ''
        return row[idx]

    results = []
    for r in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[r]
        if not row:
            continue
        address = safe_str(cell_val(row, i_address))
        asset = safe_str(cell_val(row, i_asset))
        if not address and not asset:
            continue
        results.append({
            'user_id': safe_str(cell_val(row, i_user_id)),
            'currency': asset,
            'network': safe_str(cell_val(row, i_network)),
            'amount': safe_float(cell_val(row, i_amount)),
            'address': address,
            'date': safe_str(cell_val(row, i_date)),
            'business_type': safe_str(cell_val(row, i_business_type)),
            'source': safe_str(cell_val(row, i_source)),
            'usdt_value': safe_float(cell_val(row, i_usdt)),
        })

    return results


def _parse_uol_binance_pay_tab(raw_rows):
    """
    Parse the Binance Pay tab from a UOL workbook.
    Columns: User ID, Transaction wallet ID, Merchant Name,
    Counterparty Wallet ID, Counterparty Binance ID, Order ID,
    Transaction ID, Transaction Type, Currency, Amount, Transaction Time.
    """
    if not raw_rows:
        return []

    header_row_idx = -1
    header_map = {}

    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_cp_binance_id = False
        has_transaction_id = False
        for j, cell in enumerate(row):
            cv = safe_str(cell).lower()
            if cv == 'counterparty binance id':
                has_cp_binance_id = True
            if cv == 'transaction id':
                has_transaction_id = True
        if has_cp_binance_id and has_transaction_id:
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return []

    def col_idx(possible_names):
        for name in possible_names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return -1

    i_user_id = col_idx(['User ID'])
    i_tx_wallet_id = col_idx(['Transaction wallet ID', 'Transaction Wallet ID'])
    i_merchant_name = col_idx(['Merchant Name'])
    i_cp_wallet_id = col_idx(['Counterparty Wallet ID'])
    i_cp_binance_id = col_idx(['Counterparty Binance ID'])
    i_order_id = col_idx(['Order ID'])
    i_transaction_id = col_idx(['Transaction ID'])
    i_transaction_type = col_idx(['Transaction Type'])
    i_currency = col_idx(['Currency'])
    i_amount = col_idx(['Amount'])
    i_transaction_time = col_idx(['Transaction Time'])

    def cv(row, idx):
        if idx < 0 or idx >= len(row):
            return ''
        return row[idx]

    results = []
    for r in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[r]
        if not row:
            continue
        cp_id = safe_str(cv(row, i_cp_binance_id))
        currency = safe_str(cv(row, i_currency))
        if not cp_id and not currency:
            continue
        results.append({
            'user_id': safe_str(cv(row, i_user_id)),
            'transaction_wallet_id': safe_str(cv(row, i_tx_wallet_id)),
            'merchant_name': safe_str(cv(row, i_merchant_name)),
            'counterparty_wallet_id': safe_str(cv(row, i_cp_wallet_id)),
            'counterparty_binance_id': cp_id,
            'order_id': safe_str(cv(row, i_order_id)),
            'transaction_id': safe_str(cv(row, i_transaction_id)),
            'transaction_type': safe_str(cv(row, i_transaction_type)),
            'currency': currency,
            'amount': safe_float(cv(row, i_amount)),
            'transaction_time': safe_str(cv(row, i_transaction_time)),
        })

    return results


def _parse_uol_p2p_tab(raw_rows):
    """
    Parse the P2P tab from a UOL workbook.
    Columns: Target UID, Order ID, Ad number, Buy or Sell, Crypto, Amount,
    Fiat Currency, Total Amount, Unit Price, Taker ID, Ad Publisher ID,
    Create Time, Status, Payment method, Client, Payment Time, Release time,
    Update time.
    """
    if not raw_rows:
        return []

    header_row_idx = -1
    header_map = {}

    for i, row in enumerate(raw_rows):
        if not row:
            continue
        has_taker_id = False
        has_target_uid = False
        for j, cell in enumerate(row):
            cv = safe_str(cell).lower()
            if cv in ('taker id', 'take id'):
                has_taker_id = True
            if cv == 'target uid':
                has_target_uid = True
        if has_taker_id and has_target_uid:
            header_row_idx = i
            for k, cell in enumerate(row):
                header_name = safe_str(cell)
                if header_name:
                    header_map[header_name.lower()] = k
            break

    if header_row_idx < 0:
        return []

    def col_idx(possible_names):
        for name in possible_names:
            idx = header_map.get(name.lower())
            if idx is not None:
                return idx
        return -1

    i_target_uid = col_idx(['Target UID'])
    i_order_id = col_idx(['Order ID'])
    i_ad_number = col_idx(['Ad number', 'Ad Number'])
    i_buy_sell = col_idx(['Buy or Sell'])
    i_crypto = col_idx(['Crypto'])
    i_amount = col_idx(['Amount'])
    i_fiat_currency = col_idx(['Fiat Currency'])
    i_total_amount = col_idx(['Total Amount'])
    i_unit_price = col_idx(['Unit Price'])
    i_taker_id = col_idx(['Taker ID', 'Take ID'])
    i_ad_publisher_id = col_idx(['Ad Publisher ID', 'Ad publisher ID'])
    i_create_time = col_idx(['Create Time'])
    i_status = col_idx(['Status'])
    i_payment_method = col_idx(['Payment method', 'Payment Method'])
    i_client = col_idx(['Client'])
    i_payment_time = col_idx(['Payment Time'])
    i_release_time = col_idx(['Release time', 'Release Time'])
    i_update_time = col_idx(['Update time', 'Update Time'])

    def cv(row, idx):
        if idx < 0 or idx >= len(row):
            return ''
        return row[idx]

    results = []
    for r in range(header_row_idx + 1, len(raw_rows)):
        row = raw_rows[r]
        if not row:
            continue
        taker = safe_str(cv(row, i_taker_id))
        target = safe_str(cv(row, i_target_uid))
        if not taker and not target:
            continue
        results.append({
            'target_uid': target,
            'order_id': safe_str(cv(row, i_order_id)),
            'ad_number': safe_str(cv(row, i_ad_number)),
            'buy_or_sell': safe_str(cv(row, i_buy_sell)),
            'crypto': safe_str(cv(row, i_crypto)),
            'amount': safe_float(cv(row, i_amount)),
            'fiat_currency': safe_str(cv(row, i_fiat_currency)),
            'total_amount': safe_float(cv(row, i_total_amount)),
            'unit_price': safe_float(cv(row, i_unit_price)),
            'taker_id': taker,
            'ad_publisher_id': safe_str(cv(row, i_ad_publisher_id)),
            'create_time': safe_str(cv(row, i_create_time)),
            'status': safe_str(cv(row, i_status)),
            'payment_method': safe_str(cv(row, i_payment_method)),
            'client': safe_str(cv(row, i_client)),
            'payment_time': safe_str(cv(row, i_payment_time)),
            'release_time': safe_str(cv(row, i_release_time)),
            'update_time': safe_str(cv(row, i_update_time)),
        })

    return results