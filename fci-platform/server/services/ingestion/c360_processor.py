"""
C360 Processing Pipeline.

Async wrapper around the synchronous ICR toolkit processors.
Parses uploaded C360/UOL spreadsheets, auto-populates processing
parameters from the data itself, runs all 10 processors, and returns
structured output for writing into the case document.

The toolkit code in services/icr/ is synchronous and CPU-bound.
All calls are wrapped with asyncio.to_thread() to avoid blocking
the FastAPI event loop.

Key design decisions:
  - Every processor section appears in the output markdown, even when
    the data source was not uploaded. The investigation AI must never
    have to guess whether data is missing or simply not loaded.
  - Auto-population is server-side only. There is no form for the
    investigator to fill in — parameters are derived from the uploaded
    spreadsheets, with defaults for anything that can't be detected.
  - Wallet addresses and the Elliptic CSV are extracted as separate
    return values for downstream processing (Elliptic API call, CSV
    download button).
"""

import asyncio
import io
import logging
import traceback

import openpyxl

from ..icr.parser import parse_uploaded_file
from ..icr.file_detector import classify_files
from ..icr.processor_registry import discover_processors
from ..icr.health_check import run_health_check
from ..icr.utils import safe_str, safe_int

logger = logging.getLogger(__name__)


# ── Helpers ───────────────────────────────────────────────────────


def _get_ci(row, key):
    """
    Case-insensitive row column lookup.

    C360 exports sometimes change column header casing or add
    whitespace. This ensures we find the column regardless.
    """
    if key in row:
        return row[key]
    key_lower = key.lower().strip()
    for k, v in row.items():
        if safe_str(k).lower().strip() == key_lower:
            return v
    return None


# Human-readable messages for processors whose data was not uploaded.
# Keyed by processor.id — used when should_run() returns False.
_NOT_UPLOADED_MESSAGES = {
    'tx_summary': (
        'No transaction summary data (IO Summary, IO Count, or Trade Summary) '
        'was uploaded for this case.'
    ),
    'user_profile': (
        'No extended user profile data (User Info Extended or Fiat Partners) '
        'was uploaded for this case.'
    ),
    'privacy_coin': 'No privacy coin data was uploaded for this case.',
    'counterparty': (
        'No counterparty data (Internal Transfers, Binance Pay, or P2P) '
        'was uploaded for this case.'
    ),
    'device': (
        'No device analysis data (Device Main or Device Summary) '
        'was uploaded for this case.'
    ),
    'elliptic': (
        'No address data (Top 10 by Value or Exposed Addresses) was uploaded '
        'for this case. No wallet addresses were extracted.'
    ),
    'fiat': 'No failed fiat withdrawal data was uploaded for this case.',
    'ctm': 'No CTM alert data was uploaded for this case.',
    'ftm': 'No FTM alert data was uploaded for this case.',
    'blocks': 'No account block data was uploaded for this case.',
}


# ── UID Detection ─────────────────────────────────────────────────


def _extract_subject_uid(detected, uol_data):
    """
    Extract subject UID from uploaded data.

    Priority chain (matches the existing toolkit):
      1. UOL Customer Information -> user_id
      2. C360 Device Summary -> user_id
      3. C360 Top 10 by Value -> major_user_id

    Returns:
        UID string, or empty string if no UID could be detected.
    """
    # Priority 1: UOL
    if uol_data and uol_data.get('customer_info'):
        uid = uol_data['customer_info'].get('user_id')
        if uid:
            return safe_str(uid)

    # Priority 2: Device Summary
    if 'DEVICE_SUM' in detected and detected['DEVICE_SUM']['rows']:
        uid = detected['DEVICE_SUM']['rows'][0].get('user_id', '')
        if uid:
            return safe_str(uid)

    # Priority 3: Top 10 by Value
    if 'ADDR_VALUE' in detected and detected['ADDR_VALUE']['rows']:
        uid = detected['ADDR_VALUE']['rows'][0].get('major_user_id', '')
        if uid:
            return safe_str(uid)

    return ''


# ── Auto-Population ───────────────────────────────────────────────


def _auto_populate_params(detected, uol_data, params):
    """
    Build processing params by auto-populating from uploaded data.

    Merges auto-detected values with any values already in params
    (e.g. subject_uid from the case document). Explicit params take
    precedence over auto-detected values.

    Sources (in priority order for nationality/residence):
      - UOL Customer Info -> nationality, residence (both set to
        UOL nationality — UOL has no separate residence field),
        account_type from auth_type
      - C360 USER_INFO -> nationality, residence, account_type
        (overrides UOL values when both are present)
      - USER_INFO_FIAT_PARTNERS -> ip_locations, devices_used, sys_langs
      - CP_SUMMARY -> it_count
      - DEVICE_LINK_SUMMARY -> device_link_count

    Returns:
        (merged_params, auto_populated) — merged dict for processors,
        and the raw auto-populated dict for metadata/debugging.
    """
    auto = {}
    uol_was_source = False

    # ── UOL Customer Information (Priority 1) ─────────────────
    if uol_data and uol_data.get('customer_info'):
        ci = uol_data['customer_info']
        if ci.get('user_id'):
            uol_was_source = True

            nat = ci.get('nationality')
            if nat and nat != '(null)':
                auto['nationality'] = nat
                # UOL has no separate residence field — use nationality
                auto['residence'] = nat

            auth = (ci.get('auth_type') or '').lower()
            if auth == 'personal':
                auto['account_type'] = 'individual'
            elif auth in ('enterprise', 'corporate', 'company'):
                auto['account_type'] = 'corporate'

    # ── C360 USER_INFO (Priority 2 — overrides UOL) ──────────
    if 'USER_INFO' in detected and detected['USER_INFO']['rows']:
        ui = detected['USER_INFO']['rows'][0]

        if uol_was_source:
            # UOL was primary, but C360 has more accurate nat/res fields
            nat = safe_str(ui.get('Nationality', ''))
            res = safe_str(ui.get('Residency', ''))
            if res and res != '(null)':
                auto['residence'] = res
            if nat and nat != '(null)':
                auto['nationality'] = nat
        else:
            # C360 is the only source
            nat = safe_str(ui.get('Nationality', ''))
            res = safe_str(ui.get('Residency', ''))
            user_type = safe_str(ui.get('User Type', '')).lower()

            if nat and nat != '(null)':
                auto['nationality'] = nat
            if res and res != '(null)':
                auto['residence'] = res
            if user_type == 'retail':
                auto['account_type'] = 'individual'
            elif user_type:
                auto['account_type'] = 'corporate'

    # ── Fiat Partners -> devices, IPs, languages ──────────────
    if 'USER_INFO_FIAT_PARTNERS' in detected and detected['USER_INFO_FIAT_PARTNERS']['rows']:
        fp = detected['USER_INFO_FIAT_PARTNERS']['rows'][0]
        for col, key in [
            ('IP location Used', 'ip_locations'),
            ('No. of Device Used', 'devices_used'),
            ('System Language used', 'sys_langs'),
        ]:
            val = _get_ci(fp, col)
            if val is not None:
                v = safe_int(val)
                if v > 0:
                    auto[key] = v

    # ── CP Summary -> internal transfer count ─────────────────
    if 'CP_SUMMARY' in detected and detected['CP_SUMMARY']['rows']:
        cp = detected['CP_SUMMARY']['rows'][0]
        val = _get_ci(cp, 'Internal CP User Count')
        if val is not None:
            v = safe_int(val)
            if v >= 0:
                auto['it_count'] = v

    # ── Device Link Summary -> device link count ──────────────
    if 'DEVICE_LINK_SUMMARY' in detected and detected['DEVICE_LINK_SUMMARY']['rows']:
        dl = detected['DEVICE_LINK_SUMMARY']['rows'][0]
        val = _get_ci(dl, 'Device Linked User Count')
        if val is not None:
            v = safe_int(val)
            if v >= 0:
                auto['device_link_count'] = v

    # ── Default account_type ──────────────────────────────────
    if 'account_type' not in auto:
        auto['account_type'] = 'individual'

    # ── Merge: explicit params override auto-populated values ─
    merged = dict(auto)
    for k, v in params.items():
        if v is not None and v != '':
            merged[k] = v

    return merged, auto


# ── Required Field Check ──────────────────────────────────────────


def _check_required_fields(processor, params, available):
    """
    Check if a processor's required fields are satisfied.

    Respects the 'when' condition on each field requirement —
    some fields are only required when certain file types are present
    (e.g. it_count is only required when IT spreadsheet is uploaded).

    Returns:
        List of missing field names, or empty list if all satisfied.
    """
    missing = []
    for req in processor.required_fields:
        field = req['field']

        # Evaluate 'when' condition
        when = req.get('when', 'always')
        if when != 'always' and callable(when):
            if not when(available, params):
                continue
        elif when != 'always':
            continue

        value = params.get(field)
        if not value:
            missing.append(field)
        elif isinstance(value, (int, float)) and value == 0:
            missing.append(field)

    return missing


# ── UOL Summary ───────────────────────────────────────────────────


def _build_uol_info(uol_data):
    """Build UOL summary statistics for metadata."""
    if not uol_data:
        return None

    fw = uol_data.get('fiat_withdrawals', [])
    fw_failed = sum(1 for r in fw if r.get('status', '').upper() != 'SUCCESS')

    return {
        'sheet_names': uol_data.get('sheet_names', []),
        'has_customer_info': uol_data.get('customer_info') is not None,
        'fiat_withdrawal_count': len(fw),
        'fiat_withdrawal_failed': fw_failed,
        'crypto_withdrawal_count': len(uol_data.get('crypto_withdrawals', [])),
        'crypto_deposit_count': len(uol_data.get('crypto_deposits', [])),
        'binance_pay_count': len(uol_data.get('binance_pay', [])),
        'p2p_count': len(uol_data.get('p2p_transactions', [])),
    }


# ── Quick Extraction (synchronous) ────────────────────────────────


def _is_uol_file(filename, file_bytes):
    """
    Lightweight UOL detection — reads only sheet names, no cell data.

    Returns True if the file is a UOL workbook (multi-sheet with
    'Customer Information' tab). Used to skip UOL during quick
    extraction since parsing all UOL tabs is slow.
    """
    if not filename.lower().endswith(('.xlsx', '.xls')):
        return False
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True,
        )
        try:
            names = [n.lower().strip() for n in wb.sheetnames]
            return 'customer information' in names and len(wb.sheetnames) >= 3
        finally:
            wb.close()
    except Exception:
        return False


def _extract_quick_info(files_bytes):
    """
    Fast extraction of UID from uploaded files.

    Skips UOL workbooks (slow to parse) — only parses lightweight
    C360 CSVs and single-sheet Excel files. UID comes from
    DEVICE_SUM or ADDR_VALUE fallback chain. User info from UOL
    populates later when the full background pipeline completes.

    Returns:
        Dict with file_uid, detected_file_types.
    """
    parsed_files = []
    for name, content in files_bytes:
        # Skip UOL workbooks — too slow for quick extraction
        if _is_uol_file(name, content):
            continue
        try:
            results = parse_uploaded_file(name, content)
            parsed_files.extend(results)
        except Exception as e:
            logger.warning('Quick parse failed for %s: %s', name, e)
            parsed_files.append({
                'filename': name,
                'error': str(e),
                'headers': [],
                'rows': [],
            })

    # Classify (no UOL data since we skipped it)
    detected, undetected, uol_data = classify_files(parsed_files)

    # Extract UID from C360 files only (DEVICE_SUM or ADDR_VALUE)
    file_uid = _extract_subject_uid(detected, None)

    return {
        'file_uid': file_uid or '',
        'user_info': None,
        'detected_file_types': list(detected.keys()),
    }


async def extract_quick_info(files_bytes):
    """Async wrapper for quick extraction. Offloads to thread pool."""
    return await asyncio.to_thread(_extract_quick_info, files_bytes)


# ── Sync Pipeline ─────────────────────────────────────────────────


def _run_sync_pipeline(files_bytes, params):
    """
    Synchronous C360 processing pipeline.

    Steps:
      1. Parse all uploaded files
      2. Classify into known C360 types + UOL
      3. Extract subject UID (fallback chain)
      4. Auto-populate processing parameters from data
      5. Run health checks (column changes, missing data)
      6. Run all processors — emit explicit "not uploaded" for missing ones
      7. Extract wallet addresses and CSV from elliptic processor

    Args:
        files_bytes: List of (filename, bytes) tuples.
        params: Dict — may contain subject_uid from the case document.

    Returns:
        Dict with all processing results. See process_c360_files docstring.
    """
    # ── 1. Parse ──────────────────────────────────────────────
    parsed_files = []
    for name, content in files_bytes:
        try:
            results = parse_uploaded_file(name, content)
            parsed_files.extend(results)
        except Exception as e:
            logger.warning('Failed to parse %s: %s', name, e)
            parsed_files.append({
                'filename': name,
                'error': str(e),
                'headers': [],
                'rows': [],
            })

    # ── 2. Classify ───────────────────────────────────────────
    detected, undetected, uol_data = classify_files(parsed_files)

    # ── 3. Extract UID ────────────────────────────────────────
    file_uid = _extract_subject_uid(detected, uol_data)
    subject_uid = params.get('subject_uid') or file_uid

    # ── 4. Auto-populate params ───────────────────────────────
    params['subject_uid'] = subject_uid
    merged_params, auto_populated = _auto_populate_params(
        detected, uol_data, params
    )

    # ── 5. Health check ───────────────────────────────────────
    warnings = run_health_check(detected, undetected, uol_data)

    # ── 6. Discover processors and build context ──────────────
    processors = discover_processors()

    user_info = None
    if 'USER_INFO' in detected and detected['USER_INFO']['rows']:
        user_info = detected['USER_INFO']['rows'][0]

    context = {
        'subject_uid': subject_uid,
        'uol_data': uol_data or {},
        'user_info': user_info,
    }

    # Available file types — includes virtual UOL types
    available = set(detected.keys())
    if uol_data and uol_data.get('fiat_withdrawals'):
        available.add('UOL_FIAT_WITHDRAWALS')

    # ── 7. Run all processors ─────────────────────────────────
    output_parts = []
    processor_outputs = {}
    wallet_addresses = []
    csv_content = None
    csv_filename = None

    for processor in processors:
        # Case A: Data not uploaded — processor can't run
        if not processor.should_run(available, merged_params):
            message = _NOT_UPLOADED_MESSAGES.get(
                processor.id,
                'No data was uploaded for this section.',
            )
            output_parts.append(
                '### {}\n\n{}'.format(processor.label, message)
            )
            processor_outputs[processor.id] = {
                'label': processor.label,
                'content': message,
                'has_data': False,
                'skipped': True,
                'error': None,
            }
            continue

        # Case B: Data uploaded but required params missing
        missing_fields = _check_required_fields(
            processor, merged_params, available
        )
        if missing_fields:
            field_list = ', '.join(missing_fields)
            output_parts.append(
                '### {}\n\nData files were detected but processing could '
                'not complete. Missing required parameters that could not '
                'be auto-populated: {}.'.format(processor.label, field_list)
            )
            processor_outputs[processor.id] = {
                'label': processor.label,
                'content': 'Data files were detected but processing could '
                           'not complete. Missing required parameters that '
                           'could not be auto-populated: {}.'.format(field_list),
                'has_data': False,
                'skipped': False,
                'error': 'Missing fields: {}'.format(field_list),
            }
            continue

        # Case C: Run the processor
        try:
            result = processor.process(detected, merged_params, context)

            if result.has_data:
                output_parts.append(
                    '### {}\n\n{}'.format(processor.label, result.content)
                )
            else:
                # Processor ran but found nothing — still include section
                output_parts.append(
                    '### {}\n\n{}'.format(
                        processor.label,
                        result.content or 'Processing completed but no findings.',
                    )
                )

            processor_outputs[processor.id] = {
                'label': processor.label,
                'content': result.content,
                'has_data': result.has_data,
                'skipped': False,
                'error': None,
            }

            # Extract wallet addresses and CSV from elliptic processor
            if processor.id == 'elliptic' and result.metadata:
                address_map = result.metadata.get('address_map', {})
                wallet_addresses = [
                    {
                        'address': addr,
                        'sources': info.get('sources', []),
                    }
                    for addr, info in address_map.items()
                ]
                if result.csv_content:
                    csv_content = result.csv_content
                    csv_filename = result.csv_filename

        except Exception as e:
            logger.exception('Processor %s failed', processor.id)
            output_parts.append(
                '### {}\n\nERROR: {}\n\n```\n{}\n```'.format(
                    processor.label, str(e), traceback.format_exc(),
                )
            )
            processor_outputs[processor.id] = {
                'label': processor.label,
                'content': None,
                'has_data': False,
                'skipped': False,
                'error': str(e),
            }

    # ── Assemble final output ─────────────────────────────────
    assembled_output = '\n\n---\n\n'.join(output_parts)

    undetected_summary = [
        {
            'filename': f.get('filename', 'Unknown'),
            'row_count': f.get('row_count', 0),
            'col_count': f.get('col_count', 0),
        }
        for f in (undetected or [])
    ]

    # Extract customer_info from UOL for user_info display
    uol_customer_info = None
    if uol_data and uol_data.get('customer_info'):
        uol_customer_info = uol_data['customer_info']

    # Build UOL raw data for cross-referencing (address manager, UID search)
    uol_raw_data = None
    if uol_data:
        uol_raw_data = {
            'crypto_withdrawals': uol_data.get('crypto_withdrawals', []),
            'crypto_deposits': uol_data.get('crypto_deposits', []),
            'binance_pay': uol_data.get('binance_pay', []),
            'p2p_transactions': uol_data.get('p2p_transactions', []),
            'fiat_withdrawals': uol_data.get('fiat_withdrawals', []),
            'fiat_deposits': uol_data.get('fiat_deposits', []),
        }

    return {
        'output': assembled_output,
        'processor_outputs': processor_outputs,
        'wallet_addresses': wallet_addresses,
        'csv_content': csv_content,
        'csv_filename': csv_filename,
        'subject_uid': subject_uid,
        'file_uid': file_uid or '',
        'detected_file_types': list(detected.keys()),
        'undetected_files': undetected_summary,
        'warnings': [
            {
                'level': w['level'],
                'source': w.get('source', ''),
                'message': w['message'],
            }
            for w in warnings
        ],
        'auto_populate': auto_populated,
        'uol_info': _build_uol_info(uol_data),
        'uol_customer_info': uol_customer_info,
        'uol_raw_data': uol_raw_data,
    }


# ── Async Entry Point ─────────────────────────────────────────────


async def process_c360_files(files_bytes: list, params: dict) -> dict:
    """
    Process uploaded C360/UOL files through the toolkit pipeline.

    Offloads the synchronous, CPU-bound toolkit code to a thread
    pool to avoid blocking the FastAPI event loop.

    Args:
        files_bytes: List of (filename, bytes) tuples.
                     File bytes MUST be read in the endpoint handler
                     before spawning any background task — UploadFile
                     objects are closed by the time background tasks
                     execute.
        params: Dict — may contain subject_uid from the case document,
                plus any manual overrides. Auto-populated values will
                fill gaps but not override explicit params.

    Returns:
        Dict with keys:
            output              Assembled markdown with all sections.
            wallet_addresses    List of {address, sources} dicts.
            csv_content         Elliptic batch CSV string, or None.
            csv_filename        Elliptic CSV filename, or None.
            subject_uid         Confirmed subject UID (from params or
                                auto-detected from files).
            detected_file_types List of detected file type keys.
            undetected_files    List of unrecognised file summaries.
            warnings            List of {level, source, message} dicts
                                from the health check system.
            auto_populate       Dict of values that were auto-derived
                                from the uploaded data.
            uol_info            UOL summary statistics, or None.
    """
    return await asyncio.to_thread(_run_sync_pipeline, files_bytes, params)
